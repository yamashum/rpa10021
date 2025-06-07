# rpa.py - Minimal RPA engine implementing basic actions

import json
import os
import logging
import time
from typing import Any, Dict, List

logger = logging.getLogger(__name__)
try:
    import pyautogui
except ImportError:  # pragma: no cover - optional dependency
    pyautogui = None

try:
    import openpyxl
except ImportError:  # pragma: no cover - optional dependency
    openpyxl = None


class Step:
    """Represents a single action step."""

    def __init__(self, action: str, params: Dict[str, Any]):
        self.action = action
        self.params = params

    def execute(self) -> None:
        logger.info("Executing action: %s", self.action)
        if self.action == "click":
            self._click()
        elif self.action == "input":
            self._input()
        elif self.action == "screenshot":
            self._screenshot()
        elif self.action == "file_copy":
            self._file_copy()
        elif self.action == "excel_write":
            self._excel_write()
        elif self.action == "if":
            self._if()
        elif self.action == "for":
            self._for()
        elif self.action == "wait":
            self._wait()
        elif self.action == "notify":
            self._notify()
        else:
            raise ValueError(f"Unknown action: {self.action}")

    def _click(self) -> None:
        if not pyautogui:
            raise RuntimeError("pyautogui is required for click action")
        x = self.params.get("x")
        y = self.params.get("y")
        logger.info("Click at (%s, %s)", x, y)
        pyautogui.click(x, y)

    def _input(self) -> None:
        if not pyautogui:
            raise RuntimeError("pyautogui is required for input action")
        text = self.params.get("text", "")
        logger.info("Input text: %s", text)
        pyautogui.write(text)

    def _screenshot(self) -> None:
        if not pyautogui:
            raise RuntimeError("pyautogui is required for screenshot action")
        path = self.params.get("path", "screenshot.png")
        logger.info("Save screenshot to %s", path)
        image = pyautogui.screenshot()
        image.save(path)

    def _file_copy(self) -> None:
        src = self.params.get("src")
        dst = self.params.get("dst")
        if not src or not dst:
            raise ValueError("file_copy requires src and dst")
        logger.info("Copying %s to %s", src, dst)
        with open(src, "rb") as fsrc:
            with open(dst, "wb") as fdst:
                fdst.write(fsrc.read())

    def _excel_write(self) -> None:
        if not openpyxl:
            raise RuntimeError("openpyxl is required for excel_write action")
        path = self.params.get("path")
        cell = self.params.get("cell")
        value = self.params.get("value")
        if not path or not cell:
            raise ValueError("excel_write requires path and cell")
        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
        else:
            wb = openpyxl.Workbook()
        ws = wb.active
        logger.info("Writing %s to %s in %s", value, cell, path)
        ws[cell] = value
        wb.save(path)

    def _if(self) -> None:
        condition = self.params.get("condition")
        steps = self.params.get("steps", [])
        if condition:
            for step in steps:
                step.execute()

    def _for(self) -> None:
        count = int(self.params.get("count", 0))
        steps = self.params.get("steps", [])
        for _ in range(count):
            for step in steps:
                step.execute()

    def _wait(self) -> None:
        seconds = int(self.params.get("seconds", 0))
        logger.info("Waiting for %s seconds", seconds)
        if seconds > 0:
            time.sleep(seconds)

    def _notify(self) -> None:
        message = self.params.get("message", "")
        logger.info("Notify: %s", message)
        email = self.params.get("email")
        if email:
            self._send_email(email, message)

    def _send_email(self, to_addr: str, body: str) -> None:
        import smtplib
        from email.message import EmailMessage

        server = self.params.get("smtp_server", "localhost")
        subject = self.params.get("subject", "Notification")
        from_addr = self.params.get("from_addr", "noreply@example.com")
        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = from_addr
        msg["To"] = to_addr
        msg.set_content(body)
        with smtplib.SMTP(server) as smtp:
            smtp.send_message(msg)


class Workflow:
    """Loads and executes steps from a workflow JSON file."""

    def __init__(self, steps: List[Step]):
        self.steps = steps

    @classmethod
    def load(cls, path: str) -> "Workflow":
        with open(path, "r", encoding="utf-8") as f:
            raw_steps = json.load(f)

        def build_step(data: Dict[str, Any]) -> Step:
            params = data.get("params", {})
            if "steps" in params:
                params["steps"] = [build_step(s) for s in params["steps"]]
            return Step(data["action"], params)

        steps = [build_step(step) for step in raw_steps]
        return cls(steps)

    def run(self) -> None:
        logger.info("Starting workflow with %d steps", len(self.steps))
        for step in self.steps:
            step.execute()
        logger.info("Workflow finished")


if __name__ == "__main__":  # pragma: no cover - manual execution
    import argparse

    parser = argparse.ArgumentParser(description="Run RPA workflow")
    parser.add_argument("workflow", help="Path to workflow JSON file")
    args = parser.parse_args()
    workflow = Workflow.load(args.workflow)
    workflow.run()
